# - List all files in current directory and select one to read from
import os

from pandas import ExcelWriter


def list_files():
    file_data = []
    i = 1
    for file in os.listdir("./"):
        if file.endswith(".xlsx"):
            print(str(i) + '. ', file)
            file_data.append(file)
            i = i + 1
    return file_data


file_data = list_files()
file_num = int(input("please enter no. of files: "))
file = file_data[file_num - 1]
print(file)
city = file[7:9]
# - Read from selected file and choose selection of range
import pandas as pd

sourceSheet = pd.read_excel(file, sheet_name='New Hire')
sourceSheet = sourceSheet.iloc[2:, :34]
sourceSheet.dropna(subset=["SAP ID(员工的唯一标识)"], inplace=True)
print(sourceSheet)

targetSheet = pd.read_excel("格拉默新test.xlsx", sheet_name='New Hire')
configSheet = pd.read_excel("格拉默新test.xlsx", sheet_name='城市客户名称配置表')

configSheet=configSheet.iloc[:,:2]
company_name = configSheet[configSheet["城市"]==city]["客户名称"]
sourceSheet["客户名称"]=company_name[1]
targetSheet.dropna(subset=["证件姓名及银行开户名"], inplace=True)
print(targetSheet)
count_row = targetSheet.shape[0]  # Gives number of rows
count_col = targetSheet.shape[1]  # Gives number of columns
print("start inserting from row "+str(count_row))

# Paste the dataframe to destination Excel
import pandas
from openpyxl import load_workbook
"""
book = load_workbook('格拉默新test.xlsx')
writer = pandas.ExcelWriter('格拉默新test.xlsx', engine='openpyxl')
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
sourceSheet.to_excel(writer, startrow = writer.sheets['New Hire'].max_row, startcol=1, index=False, engine='openpyxl')
writer.save()
"""
df1 = sourceSheet
df1.to_excel("output.xlsx")

print("success")